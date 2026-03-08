import io
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

# Minimum similarity ratio (0-1) to accept a fuzzy match
FUZZY_THRESHOLD = 0.80


def normalize(name: str) -> str:
    """Normalize a name: uppercase, dot→space, strip extra whitespace."""
    return " ".join(name.upper().replace(".", " ").split())


def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()


def build_lookup(extracted_rows: list[dict]) -> dict[str, float]:
    """
    Build a lookup dict from extracted image data.

    Keys are normalized "LASTNAME FIRSTNAME", values are float amounts.
    If duplicate keys exist (same name in multiple images), last one wins.
    """
    lookup: dict[str, float] = {}
    for row in extracted_rows:
        last = row.get("last_name", "").strip()
        first = row.get("first_name", "").strip()
        amount = row.get("amount")
        if last and first and amount is not None:
            key = normalize(f"{last} {first}")
            lookup[key] = float(amount)
    return lookup


def load_xlsx(file) -> tuple:
    """
    Load an xlsx file and detect the Full Name and Amount column indices.

    Args:
        file: file-like object or path

    Returns:
        (workbook, sheet, name_col_idx, amount_col_idx, header_row_idx)
        Column indices are 1-based (openpyxl convention).
    """
    wb = load_workbook(file)
    ws = wb.active

    name_col = None
    amount_col = None
    header_row = None

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value.strip().lower()
                if "full name" in val or val == "name":
                    name_col = cell.column
                    header_row = cell.row
                elif "amount" in val:
                    amount_col = cell.column

        if name_col and amount_col:
            break

    if name_col is None or amount_col is None:
        raise ValueError(
            f"Could not find required columns. "
            f"Full Name column {'found' if name_col else 'NOT found'}, "
            f"Amount column {'found' if amount_col else 'NOT found'}. "
            f"Please ensure the sheet has headers named 'Full Name' and 'Amount'."
        )

    return wb, ws, name_col, amount_col, header_row


def match_and_fill(
    ws,
    name_col: int,
    amount_col: int,
    header_row: int,
    lookup: dict[str, float],
    threshold: float = FUZZY_THRESHOLD,
) -> tuple[list[str], list[str]]:
    """
    Fill Amount cells in the sheet using the lookup dict.

    Returns:
        (matched_names, unmatched_names)
    """
    matched = []
    unmatched = []

    for row in ws.iter_rows(min_row=header_row + 1):
        name_cell = next((cell for cell in row if cell.column == name_col), None)

        if name_cell is None:
            continue

        cell_value = name_cell.value
        if cell_value is None:
            continue

        raw_name = str(cell_value).strip()
        if not raw_name:
            continue
        key = normalize(raw_name)

        # Try exact match first, then fuzzy
        amount = lookup.get(key)
        best_match_key = None

        if amount is None:
            best_ratio = 0.0
            for lookup_key, lookup_amount in lookup.items():
                ratio = similarity(key, lookup_key)
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_match_key = lookup_key
                    amount = lookup_amount
            if best_ratio < FUZZY_THRESHOLD:
                amount = None
                best_match_key = None

        if amount is not None:
            for cell in row:
                if cell.column == amount_col:
                    cell.value = amount
                    break
            label = raw_name
            if best_match_key:
                label = f"{raw_name} → {best_match_key} (fuzzy)"
            matched.append(label)
        else:
            unmatched.append(raw_name)

    return matched, unmatched


def save_xlsx(wb: Workbook) -> bytes:
    """Save workbook to bytes for Streamlit download."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
